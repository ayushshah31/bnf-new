import time
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementNotInteractableException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
wb = load_workbook('/Users/Ayush/Desktop/bread/bread.xlsx')
ws = wb['Data']

options = Options()
options.add_argument("start-maximized")
options.add_experimental_option("excludeSwitches", ['enable-automation'])

w = webdriver.Chrome(
    executable_path=ChromeDriverManager().install(), chrome_options=options)
w.get("https://www.bankofbaroda.in/personal-banking/offers")
time.sleep(10)
for z in range(100, 150):

    l = '// *[@id = "content"]/section/div/div[3]/div[2]/div/div[1]/div[' + \
        str(z) + "]/div/h5"
    # # # print(l)
    # p = '//*[@id = "content"]/section/div/div[3]/div[2]/div/div[1]/div[' + \
    #     str(z) + ']/div/div[1]/span/text()'
    # # # print(p)//*[@id="offerDetails"]/div[2]/div/div[2]/div/div[3]/div/div[2]/div
    k = '// *[@id = "content"]/section/div/div[3]/div[2]/div/div[1]/div[' + \
        str(z) + ']/div/a/h2'
    # print(k)
    #a = '//*[@id = "ulGrabDeals"]/li[' + str(z) + ']/div/div[2]/div[3]/span[1]'
    # //*[@id = "offerDetails"]/div[15
    # //*[@id = "offerDetails"]/div[51
    # print(p)
    ab = '// *[@id = "content"]/section/div/div[3]/div[2]/div/div[1]/div[' + \
        str(z) + ']/div/div[2]/a'
    # //*[@id = "offerDetails"]/div[7]/div/div[2]/div/div[4]/div[1]/a
    # //*[@id = "offerDetails"]/div[15]/div/div[2]/div[2]/div[1]/a
    # //*[@id = "offerDetails"]/div[17]/div/div[2]/div[2]/div[1]/a
    # //*[@id = "offerDetails"]/div[5]/div/div[2]/div/div[4]/div[1]/a
    # //*[@id = "offerDetails"]/div[6]/div/div[2]/div[2]/div[1]/a
    # //*[@id = "offerDetails"]/div[7]/div/div[2]/div/div[4]/div[1]/a
    # print()
    # //*[@id = "offerDetails"]/div[5]/div/div[2]/div/div[4]/div[1]/a

    h = w.find_element(By.XPATH, l).text
    print(h)
    ws['A'+str(z)] = h
    # i = w.find_element(By.XPATH, p)
    # print(i)
    # ws['B'+str(z)] = i
    j = w.find_element(By.XPATH, k).text
    print(j)
    ws['C'+str(z)] = j
    # # #cn = w.find_element(By.XPATH, aa).text
    # print(cn)
    #ws['D'+str(z)] = cn
    cm = w.find_element(By.XPATH, ab).get_attribute('href')
    print(cm)
    ws['E'+str(z)] = cm
    # time.sleep(0.5)
    print(z)

# ∂
# //*[@id = "offerDetails"]/div[2]/div/div[2]/div/div[3]/div/div[1]/div/p
# //*[@id = "ulGrabDeals"]/li[4]/div/div[2]/div[2]/a

# //*[@id = "ulGrabDeals"]/li[18]/div/div[2]/div[1]/h4
# //*[@id = "ulGrabDeals"]/li[2]/div/div[2]/div[3]/span
# //*[@id = "ulGrabDeals"]/li[1]/div/div[2]/div[2]/a

# 5
# //*[@id = "ulGrabDeals"]/li[5]/div/div[2]/div[3]/span[2]
# //
# // *[@id="ulGrabDeals"]/li[2]/div/div[2]/div[1]/h4

wb.save('/Users/Ayush/Desktop/bread/bread.xlsx')
# wb.save('/Users/Ayush/Desktop/bread/bread.xlsx')
# ws['A1'] = 'Customer Name'
# E = w.find_element(By.XPATH, '//*[@id="continue"]')
# E.click()
# time.sleep(1)
# w.find_element(By.XPATH, '//*[@id="ap_password"]').send_keys("Hideyourself")
# time.sleep(1)
# F = w.find_element(By.XPATH, '//*[@id="signInSubmit"]')
# F.click()
# time.sleep(30)

# c = w.find_element(By.XPATH, '//*[@id="nav-xshop"]/a[2]')
# c.click()

# time.sleep(1)

# d = w.find_element(By.XPATH, '//*[@id="Insurance"]')
# d.click()
# time.sleep(1)

# e = w.find_element(By.XPATH, '//*[@id="a-autoid-0-announce"]')
# e.click()
# time.sleep(1)

# f = w.find_element(By.XPATH, '//*[@id="INSURANCE_0"]')
# f.click()
# time.sleep(1)

# ws['A1'] = 'Customer Name'
# ws['B1'] = 'Bill Amount'
# ws['C1'] = 'Due From - Due To'
# ws['D1'] = 'Policy Number'
# y = 1

# for x in range(677083750, 677083775):

#     pol_no = x
#     y += 1

#     def my_func(pol_no):
#         w.find_element(By.XPATH, '//*[@id="Policy Number"]').clear()
#         # time.sleep(1)
#         w.find_element(By.XPATH, '//*[@id="Policy Number"]').send_keys(pol_no)
#         # time.sleep(1)
#         #w.find_element(By.XPATH,'//*[@id="Email id"]').send_keys("m@gmail.com");
#         # time.sleep(2)
#         # //*[@id="a-popover-1"]/div/header/button
#         cell4 = 'D' + str(y)
#         ws[cell4] = pol_no
#         g = w.find_element(By.XPATH, '//*[@id="fetchBillActionId-announce"]')
#         g.click()
#         time.sleep(2.5)
#         try:
#             # time.sleep(2)
#             g = w.find_element(
#                 By.XPATH, '//*[@id="a-popover-1"]/div/header/button')
#             # time.sleep(1)

#             #g= WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "back_button")))
#             g.click()
#             # time.sleep(2)

#         except NoSuchElementException:
#             # time.sleep(2)
#             cell1 = 'A' + str(y)
#             h = w.find_element(
#                 By.XPATH, '//*[@id="fetch-bill-table"]/tbody/tr[2]/td[2]').text
#             ws[cell1] = h
#             # time.sleep(1)

#             cell2 = 'B' + str(y)
#             i = w.find_element(
#                 By.XPATH, '//*[@id="fetch-bill-table"]/tbody/tr[3]/td[2]').text
#             ws[cell2] = i
#             # time.sleep(1)

#             cell3 = 'C' + str(y)
#             j = w.find_element(
#                 By.XPATH, '//*[@id="fetch-bill-table"]/tbody/tr[4]/td[2]').text
#             ws[cell3] = j
#             # time.sleep(2)

#         except ElementNotInteractableException:
#             # time.sleep(2)
#             cell1 = 'A' + str(y)
#             h = w.find_element(
#                 By.XPATH, '//*[@id="fetch-bill-table"]/tbody/tr[2]/td[2]').text
#             ws[cell1] = h
#             # time.sleep(1)

#             cell2 = 'B' + str(y)
#             i = w.find_element(
#                 By.XPATH, '//*[@id="fetch-bill-table"]/tbody/tr[3]/td[2]').text
#             ws[cell2] = i
#             # time.sleep(1)

#             cell3 = 'C' + str(y)
#             j = w.find_element(
#                 By.XPATH, '//*[@id="fetch-bill-table"]/tbody/tr[4]/td[2]').text
#             ws[cell3] = j
#             # time.sleep(2)

#     my_func(pol_no)

# wb.save('C:\\Users\\malpa\\Desktop\\music.xlsx')
time.sleep(2)
